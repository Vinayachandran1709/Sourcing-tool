"""
One-time script: Load Indian VCs from the enriched spreadsheet into vc_sources.json.
Usage: python scripts/seed_vcs_from_xlsx.py path/to/indian_vcs_directory_enriched.xlsx
"""

import json
import os
import sys


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "indian_vcs_directory_enriched.xlsx"

    try:
        import openpyxl
    except ImportError:
        print("pip install openpyxl first")
        return

    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    config_path = os.path.join(os.path.dirname(__file__), "..", "config", "vc_sources.json")
    existing = []
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            existing = json.load(f)

    existing_names = {vc["name"].lower() for vc in existing if vc.get("name")}
    added = 0

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        website = ws.cell(row, 6).value
        stages = ws.cell(row, 3).value or ""
        sectors = ws.cell(row, 7).value or ""
        portfolio_count = ws.cell(row, 14).value or "0"

        if not name or not website:
            continue
        if str(website).startswith("ERROR"):
            continue
        if name.lower() in existing_names:
            continue

        existing.append(
            {
                "name": name,
                "portfolio_url": website,
                "website": website,
                "region": "India",
                "stages": stages,
                "sectors": sectors,
                "portfolio_count": str(portfolio_count),
                "source": "indianvcs_directory",
            }
        )
        existing_names.add(name.lower())
        added += 1

    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(existing, f, indent=2)

    print(f"Added {added} Indian VCs to {config_path}")
    print(f"Total VCs in config: {len(existing)}")


if __name__ == "__main__":
    main()
